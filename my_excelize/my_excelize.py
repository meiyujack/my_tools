from openpyxl import load_workbook, Workbook
from apiflask import APIFlask, Schema
from apiflask.fields import String, File
from flask import render_template, redirect, url_for, flash, send_file

import os


class Excelize:

    def __init__(self, file_address="upload/model.xlsx", sheet_name=None):
        if os.path.exists(file_address):
            self.file_address = file_address
        else:
            wb = Workbook()
            ws = wb.active
            ws["A1"] = 0
            wb.save("upload/model.xlsx")
            self.file_address = "upload/model.xlsx"
        self.workbook = load_workbook(self.file_address)
        self.sheet = (self.workbook.active
                      if not sheet_name else self.workbook[sheet_name])

    def get_basic(self):
        rows = self.sheet.max_row
        cols = self.sheet.max_column
        return rows, cols, self.sheet.columns.__next__()[0].column_letter

    def get_th(self):
        return [th[0].value for th in self.sheet.columns]

    def write_data(self, data: list):
        rows, cols, start_letter = self.get_basic()
        for c in range(cols):
            end_letter = chr(ord(start_letter) + c)
            self.sheet[end_letter + str(rows + 1)] = data[c]
        self.workbook.save(self.file_address)

    def read_data(self):
        rows, cols, start_letter = self.get_basic()
        table = []
        for r in range(2, rows + 1):
            row = []
            for c in range(cols):
                end_letter = chr(ord(start_letter) + c)
                row.append(self.sheet[end_letter + str(r)].value)
            table.append(row)
        return table


class InputFile(Schema):
    model = File(required=True)


my_excel = APIFlask(__name__)
my_excel.secret_key = "secret key"


@my_excel.get("/")
def index():
    flag = False
    if Excelize("upload/model.xlsx").get_basic()[0] > 1:
        flag = True
    return render_template("index.html", flag=flag)


@my_excel.get("/upload")
def upload():
    return render_template("upload.html")


@my_excel.post("/model")
@my_excel.input(InputFile, location="files")
def upload_model(files_data):
    with open("upload/model.xlsx", "wb") as f:
        f.write(files_data["model"].read())
    flash("模板上传成功！")
    return redirect(url_for("write_show"))


@my_excel.get("/write")
def write_show():
    if Excelize("upload/model.xlsx").get_th() == [0]:
        flash("当前没有模板，请先上传含表头的excel文件（xlsx）")
        return redirect(url_for("upload"))
    excel = Excelize("upload/model.xlsx")
    th = excel.get_th()
    return render_template("write_excel.html", th=th)


@my_excel.post("/write")
@my_excel.input(
    {
        "column" + str(k): String(required=False)
        for k in range(len(Excelize("upload/model.xlsx").get_th()))
    },
    location="form",
)
def write_post(form_data):
    excel = Excelize("upload/model.xlsx")
    content = []
    for v in form_data.values():
        content.append(v)

    excel.write_data(content)
    flash("数据写入成功！")
    return redirect(url_for("read_show"))


@my_excel.get("/read")
def read_show():
    excel = Excelize("upload/model.xlsx")
    if excel.get_basic()[0] == 1:
        flash("暂无数据，请写入~")
        return redirect(url_for("write_post"))
    return render_template("read.html",
                           th=excel.get_th(),
                           table=excel.read_data())


@my_excel.get("/download")
def download():
    return send_file(path_or_file="upload/model.xlsx",
                     download_name="工作簿.xlsx")
