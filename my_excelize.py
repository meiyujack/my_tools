from openpyxl import load_workbook


class Excelize:
    def __init__(self, file_address, sheet_name=None):
        self.file_address = file_address
        self.workbook = load_workbook(self.file_address)
        self.sheet = (
            self.workbook.active if not sheet_name else self.workbook[sheet_name]
        )

    def get_basic(self):
        rows = self.sheet.max_row
        cols = self.sheet.max_column
        return rows, cols, self.sheet.columns.__next__()[0].column_letter

    def write_data(self, data: list):
        rows, cols, start_letter = self.get_basic()
        for c in range(cols):
            end_letter = chr(ord(start_letter) + c)
            self.sheet[end_letter + str(rows + 1)] = data[c]
        self.workbook.save(self.file_address)
